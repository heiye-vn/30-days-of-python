"""
不同格式文件的处理方式，如 txt、json、yaml、csv、excel、图片、音频、视频等
"""

import configparser

# import csv
import json
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

import markdown
import openpyxl
import pandas as pd
import pdfplumber
import yaml
from openpyxl.styles import Font, PatternFill
from PIL import Image, ImageDraw, ImageEnhance, ImageFilter, ImageFont
from pypdf import PdfReader, PdfWriter
from reportlab.lib.pagesizes import A4
from reportlab.lib.pdfencrypt import encryptCanvas
from reportlab.pdfgen import canvas

""" 1 文本文件（.txt） """
# 读取
text_1 = Path("data/sample.txt").read_text(encoding="utf-8")
# print(text_1)

# 写入（覆盖）
# Path("data/output.txt").write_text("Hello World!", encoding="utf-8")

# 编码处理，常见编码：utf-8、gbk、gb2312、latin-1、uft-16
# with open("data/gbk_file.txt", "r", encoding="gbk") as f:
#     content = f.read()

"""
处理编码错误
errors 参数选项:
    'strict': 默认值，报错
    'ignore': 忽略无法解码的字节
    'replace': 替换为 ? 或 U+FFFD
    'surrogateescape': 使用代理对
    'backslashreplace': 将无法解码的字符替换为反斜杠转义序列
"""
# with open("data/sample.txt", "r", encoding="utf-8", errors="ignore") as f:
#     content = f.read()


"""
2. csv 文件
CSV 是最常见的格式化数据格式
"""
# ========== 标准库 csv ==========

# 读取 CSV
# with open("data/sample.csv", "r", encoding="utf-8") as f:
#     reader = csv.reader(f)
#     header = next(reader)  # 读取表头
#     print(f"列名：{header}")
#     for row in reader:
#         print(row)

# 以字典方式读取（⭐⭐⭐⭐⭐）
# with open("data/sample.csv", "r+", encoding="utf-8", newline="") as f:
#     # 移动指针到文件末尾，防止覆盖旧数据
#     f.seek(0, 2)
#     writer = csv.writer(f)
#     # 追加数据只需写数据行
#     writer.writerows(
#         [
#             ["麻子哥", 500, "雷仙殿", 50000],
#         ]
#     )
#
#     # 将指针重置到最开头，以便从头读取整张表（含表头）
#     f.seek(0)
#
#     reader = csv.DictReader(f)
#     for row in reader:
#         print(row)
# print(f"{row['name']} - {row['city']} - {row['salary']}")

# 写入 CSV
# data = [["name", "age", "city"], ["Eve", 27, "Guangzhou"], ["Frank", 35, "Chengdu"]]
# with open("data/output.csv", "w", encoding="utf-8", newline="") as f:
#     writer = csv.writer(f)
#     writer.writerows(data)

# ========== pandas（推荐处理大型 CSV） ==========

# 读取
# df = pd.read_csv("data/sample.csv")
# print(df.head())
# print(df.describe())

# 写入
# df.to_csv("data/output_pandas.csv", index=False, encoding="utf-8")

# 高级读取选项
# df_ = pd.read_csv(
#     "data/sample.csv",
#     usecols=["name", "salary"],  # 只读特定列
#     dtype={"salary": int},  # 指定数据类型
#     nrows=10,  # 只读前 10 列
#     skiprows=[1],  # 跳过索引为 1 的行，索引 0 是表头，不能跳过
# )
# print(df_.head())


"""
3. JSON 文件
JSON 是 Web 开发和数据交换中最常用的格式。
"""

# ========== 读取 JSON ==========

# 从文件读取: json.load()
with open("data/sample.json", "r", encoding="utf-8") as f:
    data1 = json.load(f)
# print(data1["company"])
# print(data1["employees"][0]["name"])

# 从字符串读取: json.loads()
json_str = '{"name": "Alice", "age": 28}'
data_str = json.loads(json_str)
# print(data_str)

# ========== 写入 JSON ==========
write_data = {"name": "Bob111", "skills": ["Python", "SQL", "Docker", "Ai Agent"]}

# 写入文件（格式化输出）: json.dump()
# with open("data/output.json", "w", encoding="utf-8") as f:
#     json.dump(write_data, f, ensure_ascii=False, indent=2)

# 转为字符串
json_str = json.dumps(write_data, ensure_ascii=False, indent=2)
# print(json_str)

# ========== json.dumps/dump 常用参数 ==========
# ensure_ascii=False  允许输出非 ASCII 字符（中文等）
# indent=2            缩进空格数（美化输出）
# sort_keys=True      按 key 排序
# separators=(",", ":")  紧凑输出
# default=str         自定义不可序列化对象的处理函数


# ========== 处理特殊类型 ==========
class CustomEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
            # return obj.strftime("%Y-%m-%d %H:%M:%S")
        return super().default(obj)


data2 = {"timestamp": datetime.now(), "event": "login"}
json_str2 = json.dumps(data2, cls=CustomEncoder, ensure_ascii=False)
# print(json_str2)


""" 4. XML 文件 """

# ========== 读取 XML ==========

# 从文件解析
tree = ET.parse("data/sample.xml")
root = tree.getroot()
# print(f"根元素：{root.tag}")

# 遍历元素
for book in root.findall("book"):
    bool_id = book.get("id")
    title = book.find("title").text  # noqa
    price = book.find("price").text  # noqa
    # print(f"[{bool_id}] {title} - ${price}")

# XPath 查找
# titles = root.findall(".//title")
# for t in titles:
#     print(t.text)

# 从字符串解析
xml_str = "<root><item>hello</item></root>"
root_ = ET.fromstring(xml_str)
# print(root_.tag)

# ========== 创建/修改 XML ==========

# 创建新的 XML
root2 = ET.Element("users")
# user = ET.SubElement(root2, "user", id="1")
# ET.SubElement(user, "name").text = "Alice"
# ET.SubElement(user, "age").text = "28"

# 写入文件
tree2 = ET.ElementTree(root2)
# ET.indent(tree2, space=" ")  # 美化缩进
# tree2.write("data/outpt.xml", encoding="utf-8", xml_declaration=True)

# ========== 第三方库 lxml（更强大） ==========
# pip install lxml
# from lxml import etree
# tree = etree.parse("data/sample.xml")
# 支持 XPath 完整语法、XSLT 转换、Schema 验证等


"""
5. YAML 文件
YAML 广泛用于配置文件，语法简洁，可读性强
"""

# ========== 读取 YAML ==========
with open("data/config.yaml", "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

# print(config["server"]["host"])
# print(config["database"]["driver"])

# ========== 写入 YAML ==========
data3 = {
    "app": {"name": "MyApp", "version": "2.0"},
    "features": ["auth", "logging", "cache"],
    "debug": False,
}
# with open("data/output.yaml", "w", encoding="utf-8") as f:
#     yaml.dump(data3, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

# ========== 安全注意事项 ==========
# 始终使用 yaml.safe_load() 而不是 yaml.load()
# yaml.load() 可以执行任意 Python 代码，存在安全风险


""" 6. INI/配置文件 """

# ========== 读取 INI ==========
config = configparser.ConfigParser()
config.read("data/config.ini", encoding="utf-8")

# 访问值
# app_name = config["general"]["app_name"]
# db_port = config.getint("database", "port")
# debug = config.getboolean("general", "debug")
# print(f"App: {app_name}, DB Port: {db_port}, Debug: {debug}")

# 遍历所有 sections
# for section in config.sections():
#     print(f"\n[{section}]")
#     for key, value in config[section].items():
#         print(f" {key} = {value}")

# ========== 写入 INI ==========
# config["app"] = {"name": "NewApp", "version": "1.0.0", "debug": "false"}
# config["server"] = {"host": "0.0.0.0", "port": "8080"}
# with open("data/output.ini", "w", encoding="utf-8") as f:
#     config.write(f)

# ========== 修改已有配置 ==========
# config.read("data/config.ini", encoding="utf-8")
# config["general"]["version"] = "2.0.0"
# config["new_section"] = {"key1": "value1"}
# with open("data/config.ini", "w", encoding="utf-8") as f:
#     config.write(f)


""" 7. Excel 文件（.xlsx） """

# ========== 创建/写入 Excel ==========
wb = openpyxl.Workbook()  # 创建一个新的工作簿
ws = wb.active  # 获取活动工作表, 默认是 Sheet
ws.title = "员工数据"  # 设置工作表标题

# 写入表头
headers = ["姓名", "年龄", "部门", "薪资"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, size=12)
    cell.fill = PatternFill(start_color="4472C4", fill_type="solid")

# 写入数据
data4 = [
    ["Alice", 28, "工程部", 15000],
    ["Bob", 32, "市场部", 22000],
    ["Charlie", 25, "设计部", 12000],
]
for row_idx, row_data in enumerate(data4, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 调整列宽
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 10

# wb.save("data/output.xlsx")

# ========== 读取 Excel ==========
# wb_read = openpyxl.load_workbook("data/output.xlsx")
# ws_read = wb_read.active
# for row in ws_read.iter_rows(min_row=1, values_only=True):
#     print(row)

# ========== pandas 处理 Excel（数据分析首选） ==========

# 读取
# df = pd.read_excel("data/output.xlsx", sheet_name="员工数据")
# print(df)

# 写入（支持多个 sheet）
# with pd.ExcelWriter("data/multi_sheet.xlsx", engine="openpyxl") as writer:
#     df.to_excel(writer, sheet_name="Sheet1", index=False)
#     df.describe().to_excel(writer, sheet_name="统计")


""" 8. PDF 文件 """

# ========== 读取 PDF（pypdf） ==========
# reader = PdfReader("data/sample.pdf")
# print(f"总页数：{len(reader.pages)}")
# for i, page in enumerate(reader.pages):
#     text = page.extract_text()
#     print(f"--- 第 {i + 1} 页")
#     print(text)

# ========== 合并 PDF ==========
# with PdfWriter() as writer:
#     writer.append("file.pdf")
#     writer.append("file2.pdf")
#     writer.write("merged.pdf")

# ========== 创建 PDF（reportlab） ==========

# 1. 创建普通 PDF 文件
# c = canvas.Canvas("data/output.pdf", pagesize=A4)
# c.setFont("Helvetica", 12)
# c.drawString(100, 750, "Hello, World!")
# c.drawString(100, 730, "Created with reportlab.")
# c.save()

# 2. 创建加密 PDF 文件（对 canvas 进行原地加密包装）
# raw_canvas = canvas.Canvas("data/output_encrypted.pdf", pagesize=A4)
# # 原地配置加密，设置打开密码为 "my_password"
# encryptCanvas(raw_canvas, userPassword="my_password")
# raw_canvas.setFont("Helvetica", 12)
# raw_canvas.drawString(100, 750, "This is an encrypted PDF.")
# raw_canvas.save()

# ========== PDF 转文本（pdfplumber，表格提取更强） ==========
# with pdfplumber.open("data/sample.pdf") as pdf:
#     for i, page in enumerate(pdf.pages):
#         print(f"--- 第 {i + 1} 页表格数据 --- ")
#         # 提取当前页面的所有表格，返回一个三维列表：【表格数、行数、列数】
#         tables = page.extract_tables()
#         for table in tables:
#             for row in table:
#                 print(row)


""" 9. 图片文件 """

# ========== 基本操作 ==========
img = Image.open("data/image.png")
# print(f"尺寸：{img.size}")
# print(f"模式：{img.mode}")
# print(f"格式：{img.format}")

# 调整大小
img_resized = img.resize((800, 600))
# img_resized.save("data/resized_image.png")

# 裁剪（left、upper、right、lower）
img_cropped = img.crop((100, 100, 400, 400))
# img_cropped.save("data/cropped_image.png")

# 旋转
img_rotated = img.rotate(90, expand=True)
# img_rotated.save("data/rotated_image.png")

# 转换为灰度图像
img_gray = img.convert("L")
# img_gray.save("data/grayscale_image.png", "PNG")

# ========== 滤镜 ========
img_blur = img.filter(ImageFilter.BLUR)  # 模糊滤镜
# img_blur.save("data/blurred_image.png")
img_edge = img.filter(ImageFilter.FIND_EDGES)  # 边缘检测滤镜
# img_edge.save("data/edge_image.png")
img_sharpen = img.filter(ImageFilter.SHARPEN)  # 锐化滤镜
# img_sharpen.save("data/sharpened_image.png")

# ========== 图像增强 ==========
enhancer = ImageEnhance.Brightness(img)
img_bright = enhancer.enhance(1.5)
# img_bright.save("data/bright_image.png")

# ========== 批量处理示例 ==========
# for img_path in Path("data").glob("*.png"):
#     img = Image.open(img_path)
#     img.thumbnail((200, 200))  # 保持比例的缩放
#     img.save(f"data/batch_imgs/thumb_{img_path.name}")

# ========== 常见高频补充操作 ==========

# 1. 创建全新画布
bg = Image.new("RGBA", (400, 400), color="white")  # 创建 400x400 的白色透明画布

# 2. 图像绘制与文字水印（几何图形与文字）
draw = ImageDraw.Draw(img)
# 画矩形（常用于目标检测 Bounding Box 绘制）
draw.rectangle([50, 50, 200, 200], outline="red", width=3)
# 画圆形/椭圆
draw.ellipse([250, 50, 350, 150], fill="blue")
# 绘制文字水印（可以使用默认字体或加载自定义 ttf 字体）
# font = ImageFont.truetype("arial.ttf", 20)
draw.text((10, 10), "Sample Watermark", fill="yellow")
# img.save("data/drawn_image.png")

# 3. 图像拼接/贴图（如粘贴 Logo）
# logo = Image.open("data/image.png")
# 如果 logo 带有透明通道（RGBA），需将它同时作为 mask 参数传入，以保证透明度正常
# img.paste(logo, (50, 50), mask=logo)
# img.save("data/pasted_image.png")

# 4. 图像保存优化（质量压缩与大小优化）
# 对于 JPG，可以使用 quality 调节质量（1-95），optimize=True 可以进一步压缩文件体积
# img.convert("RGB").save("data/compressed.jpg", "JPEG", quality=85, optimize=True)

""" 10. Markdown 文件 """

# ========== 基础读写 ==========
# Markdown 本质上是 UTF-8 编码的纯文本文件，可以直接用内置 read_text/write_text 进行读写
md_content = """# Python 学习指南

这是正文内容，包含一个 [超链接](https://google.com) 和 **加粗文本**。

- 列表项一
- 列表项二
"""
# 写入 .md 文件
# Path("data/output.md").write_text(md_content, encoding="utf-8")

# 读取 .md 文件
# text = Path("data/sample.md").read_text(encoding="utf-8")
# print(text)


# ========== 解析与渲染（Markdown 转 HTML） ==========
with open("data/AI_Agent_Study_Plan.md", "r", encoding="utf-8") as f:
    md_text = f.read()

# html = markdown.markdown(md_text, extensions=["tables", "fenced_code"])
# print(html)

# 带扩展名的转换
html = markdown.markdown(
    md_text,
    extensions=[
        "tables",  # 表格支持
        "fenced_code",  # 代码块
        "codehilite",  # 代码高亮
        "toc",  # 目录生成
        "nl2br",  # 换行转 br
    ],
)

# 写入 html 文件，并加入简易精美的 CSS 样式来美化表格与文字布局
css_style = """
<style>
    body {
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Helvetica, Arial, sans-serif;
        line-height: 1.6;
        color: #24292e;
        max-width: 800px;
        margin: 40px auto;
        padding: 0 20px;
    }
    table {
        border-collapse: collapse;
        width: 100%;
        margin: 20px 0;
    }
    th, td {
        border: 1px solid #dfe2e5;
        padding: 10px 14px;
        text-align: left;
    }
    th {
        background-color: #f6f8fa;
        font-weight: 600;
    }
    tr:nth-child(even) {
        background-color: #f8f9fa;
    }
</style>
"""
with open("data/output.html", "w", encoding="utf-8") as f:
    f.write(f"<html><head>{css_style}</head><body>{html}</body></html>")

# ========== 提取 Front Matter（文档/博客元数据） ==========
# 很多 Markdown 文件（例如 Hugo、Hexo 静态博客）顶部会包含 YAML 格式的元数据（Front Matter），如：
# ---
# title: "Markdown 处理教学"
# date: 2026-07-05
# tags: [Python, Markdown]
# ---
# 推荐使用 `python-frontmatter` 库来进行解析（需要安装：pip install python-frontmatter）
#
# import frontmatter
# post = frontmatter.load("data/post.md")
# print("元数据字典 (Metadata):", post.metadata)
# print("文章正文内容 (Content):", post.content)
